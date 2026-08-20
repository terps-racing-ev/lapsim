"""Track segments, spreadsheet loading, and racing-line plotting."""

from collections.abc import Iterable
from dataclasses import dataclass
from math import ceil, cos, sin
from os import PathLike
from typing import Protocol

from openpyxl import load_workbook

from utils.units import degrees_to_radians, feet_to_meters


class Segment(Protocol):
    """Geometric track segment consumed by the spatial solvers."""

    @property
    def length_m(self) -> float: ...


@dataclass(slots=True)
class Straight:
    """Straight racing-line segment."""

    length_m: float


@dataclass(slots=True)
class Curve:
    """Constant-radius curve; positive span turns counterclockwise."""

    radius_m: float
    span_rad: float

    @property
    def length_m(self) -> float:
        return abs(self.radius_m * self.span_rad)


class Track:
    """Ordered list of straight and constant-radius racing-line segments."""

    def __init__(self, file_path: str | PathLike[str]) -> None:
        self.segments = self._load_segments(file_path)

    @classmethod
    def from_segments(cls, segments: Iterable[Segment]) -> "Track":
        """Build a track directly without creating a spreadsheet."""

        track = cls.__new__(cls)
        track.segments = list(segments)
        if not track.segments:
            raise ValueError("segments cannot be empty")
        return track

    @staticmethod
    def _load_segments(file_path: str | PathLike[str]) -> list[Segment]:
        segments: list[Segment] = []
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active

        try:
            for length_ft, radius_ft, span_deg in worksheet.iter_rows(
                min_col=1,
                max_col=3,
                values_only=True,
            ):
                length_ft = float(length_ft or 0.0)
                radius_ft = float(radius_ft or 0.0)
                span_deg = float(span_deg or 0.0)

                if length_ft and radius_ft:
                    raise ValueError(
                        "A track segment cannot be both straight and curved"
                    )
                if length_ft:
                    segments.append(Straight(length_m=feet_to_meters(length_ft)))
                elif radius_ft:
                    segments.append(
                        Curve(
                            radius_m=feet_to_meters(radius_ft),
                            span_rad=degrees_to_radians(span_deg),
                        )
                    )
        finally:
            workbook.close()

        if not segments:
            raise ValueError("Track spreadsheet contains no segments")
        return segments

    @property
    def total_length_m(self) -> float:
        """Total racing-line length."""

        return sum(segment.length_m for segment in self.segments)

    def plot(self, curve_resolution_rad: float = degrees_to_radians(0.5)):
        """Plot the racing-line geometry and return its figure and axes."""

        import matplotlib.pyplot as plt

        if curve_resolution_rad <= 0:
            raise ValueError("curve_resolution_rad must be positive")

        x_m = 0.0
        y_m = 0.0
        heading_rad = 0.0
        x_coordinates_m = [x_m]
        y_coordinates_m = [y_m]

        for segment in self.segments:
            if isinstance(segment, Straight):
                x_m += segment.length_m * cos(heading_rad)
                y_m += segment.length_m * sin(heading_rad)
                x_coordinates_m.append(x_m)
                y_coordinates_m.append(y_m)
                continue

            if isinstance(segment, Curve):
                starting_x_m = x_m
                starting_y_m = y_m
                starting_heading_rad = heading_rad
                turn_direction = 1.0 if segment.span_rad > 0 else -1.0
                point_count = max(
                    1,
                    ceil(abs(segment.span_rad) / curve_resolution_rad),
                )

                for point_index in range(1, point_count + 1):
                    fraction = point_index / point_count
                    heading_rad = starting_heading_rad + segment.span_rad * fraction
                    x_m = starting_x_m + turn_direction * segment.radius_m * (
                        sin(heading_rad) - sin(starting_heading_rad)
                    )
                    y_m = starting_y_m - turn_direction * segment.radius_m * (
                        cos(heading_rad) - cos(starting_heading_rad)
                    )
                    x_coordinates_m.append(x_m)
                    y_coordinates_m.append(y_m)
                continue

            raise TypeError(f"Unsupported segment type: {type(segment).__name__}")

        figure, axes = plt.subplots()
        axes.plot(x_coordinates_m, y_coordinates_m)
        axes.scatter(
            x_coordinates_m[0],
            y_coordinates_m[0],
            label="Start/finish",
        )
        axes.set_aspect("equal", adjustable="box")
        axes.set_xlabel("x [m]")
        axes.set_ylabel("y [m]")
        axes.set_title("Track")
        axes.legend()
        axes.grid(True)
        return figure, axes


__all__ = ["Curve", "Segment", "Straight", "Track"]
