"""Track representation and spatial discretization."""

from dataclasses import dataclass
from math import ceil, cos, sin
from typing import Protocol

from openpyxl import load_workbook

from .utils.units import degrees_to_radians, feet_to_meters


class Segment(Protocol):

    @property
    def length_m(self) -> float: ...


@dataclass(slots=True)
class Straight:

    length_m: float


@dataclass(slots=True)
class Curve:

    radius_m: float
    span_rad: float

    @property
    def length_m(self) -> float:
        return abs(self.radius_m * self.span_rad)

class Track:

    segments: list[Segment]

    def __init__(self, file_path: str) -> None:
        

        self.segments = []

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active

        try:
            for length_ft, radius_ft, span_deg in worksheet.iter_rows(
                min_col=1,
                max_col=3,
                values_only=True,
            ):
                length_ft = float(length_ft or 0)
                radius_ft = float(radius_ft or 0)
                span_deg = float(span_deg or 0)

                if length_ft and radius_ft:
                    raise ValueError("A track segment cannot be both straight and curved")

                if length_ft:
                    self.segments.append(
                        Straight(length_m=feet_to_meters(length_ft))
                    )
                elif radius_ft:
                    self.segments.append(
                        Curve(
                            radius_m=feet_to_meters(radius_ft),
                            span_rad=degrees_to_radians(span_deg),
                        )
                    )
        finally:
            workbook.close()

    def plot(self, curve_resolution_rad: float = degrees_to_radians(0.5)):
        import matplotlib.pyplot as plt


        if curve_resolution_rad <= 0:
            raise ValueError("curve_resolution_rad must be positive")

        x = 0.0
        y = 0.0
        heading_rad = 0.0
        x_coordinates = [x]
        y_coordinates = [y]

        for segment in self.segments:
            if isinstance(segment, Straight):
                x += segment.length_m * cos(heading_rad)
                y += segment.length_m * sin(heading_rad)
                x_coordinates.append(x)
                y_coordinates.append(y)
                continue

            if isinstance(segment, Curve):
                starting_x = x
                starting_y = y
                starting_heading = heading_rad
                turn_direction = 1.0 if segment.span_rad > 0 else -1.0
                point_count = max(
                    1,
                    ceil(abs(segment.span_rad) / curve_resolution_rad),
                )

                for point_index in range(1, point_count + 1):
                    fraction = point_index / point_count
                    heading_rad = starting_heading + segment.span_rad * fraction
                    x = starting_x + turn_direction * segment.radius_m * (
                        sin(heading_rad) - sin(starting_heading)
                    )
                    y = starting_y - turn_direction * segment.radius_m * (
                        cos(heading_rad) - cos(starting_heading)
                    )
                    x_coordinates.append(x)
                    y_coordinates.append(y)
                continue

            raise TypeError(f"Unsupported segment type: {type(segment).__name__}")

        figure, axes = plt.subplots()
        axes.plot(x_coordinates, y_coordinates)
        axes.scatter(x_coordinates[0], y_coordinates[0], label="Start/finish")
        axes.set_aspect("equal", adjustable="box")
        axes.set_xlabel("x [m]")
        axes.set_ylabel("y [m]")
        axes.set_title("Track")
        axes.legend()
        axes.grid(True)
        return figure, axes

if __name__ == "__main__":
    track = Track('EV_MI_Endur.xlsx')
    fig, ax = track.plot()
    plt.show(block=True)

    
